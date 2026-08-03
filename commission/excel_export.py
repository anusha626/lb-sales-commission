"""Build the multi-sheet Excel report.

Sheets produced:
  - Summary           : per-SA totals, tier, commission
  - One sheet per SA  : full audit trail of contributing orders
  - Review log        : orders that needed manual attention
  - Excluded          : orders excluded by status filters
  - Settings snapshot : the rate table, tiers, SA list at run time
"""
from __future__ import annotations

from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    CommissionReport,
    HouseSalesSummary,
    OrderResult,
    SACommission,
    SAContribution,
)
from .parser import HOUSE_ACCOUNT
from .settings import AppSettings, TiersConfig

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MONEY_FMT = '"RM"#,##0.00'
_PCT_FMT = "0.00\\%"

# ---- Presentation palette (per-SA sheet) -----------------------------------
_NAVY = "1F4E79"
_GOLD = "9A761F"
_TEAL = "0F766E"
_ZEBRA = "F3F6FB"
_SUBTOTAL_FILL = "E8EEF7"
_CLR_ZEBRA = "FBF5E6"
_INK = "1F2937"
_MUTED = "6B7280"
_thin = Side(style="thin", color="D6DEEA")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Redesigned per-SA sheet: one row per order, split into Paid vs Clearance.
# Columns: A Order# B Date C Channel D Payment E Gross F Discount G Charge%
#          H Charges I Net J Share% K Commission
_SA_HEADERS = [
    "Order #", "Date", "Channel", "Payment", "Gross share", "Discount",
    "Charge %", "Charges", "Net share", "Share %", "Commission",
]
_SA_NC = len(_SA_HEADERS)  # 11 columns (A–K)
_SA_NUM_COLS = {5, 6, 8, 9, 11}  # money (Gross, Discount, Charges, Net, Commission)
_SA_PCT_COLS = {7, 10}           # Charge %, Share %
_SA_RIGHT_COLS = {5, 6, 7, 8, 9, 10, 11}  # right-aligned columns

# Live-tier helper cells (top-right, off the main table). M2 = the SA's sales
# net (excl. clearance); M3 = the whole-bracket tier rate looked up from M2.
# Commission formulas reference $M$3, so editing rows re-picks the tier.
_TIER_NET_CELL = "M2"
_TIER_RATE_CELL = "M3"
_TIER_RATE_REF = "$M$3"


def _tier_rate_formula(net_cell: str, tiers) -> str:
    """Nested-IF whole-bracket tier lookup: returns the rate % for `net_cell`."""
    st = sorted(tiers, key=lambda t: t.min_net)
    formula = f"{st[0].rate_pct}"
    for t in st[1:]:
        formula = f"IF({net_cell}>={t.min_net:.0f},{t.rate_pct},{formula})"
    return "=" + formula


def _write_header(ws: Worksheet, headers: Sequence[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _autosize(ws: Worksheet, min_w: int = 10, max_w: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            longest = max(longest, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_w, min(max_w, longest + 2)
        )


def _build_summary_sheet(
    ws: Worksheet,
    summaries: list[SACommission],
    house: HouseSalesSummary | None,
) -> None:
    ws.title = "Summary"
    headers = [
        "Sales Advisor",
        "# Orders",
        "Total Gross",
        "Total Net",
        "Avg Order",
        "Tier",
        "Tier Rate",
        "Commission (RM)",
    ]
    _write_header(ws, headers)
    for i, s in enumerate(summaries, start=2):
        ws.cell(row=i, column=1, value=s.sa_name)
        ws.cell(row=i, column=2, value=s.order_count)
        ws.cell(row=i, column=3, value=s.total_gross_sales).number_format = _MONEY_FMT
        ws.cell(row=i, column=4, value=s.total_net_sales).number_format = _MONEY_FMT
        ws.cell(row=i, column=5, value=s.avg_order_value).number_format = _MONEY_FMT
        ws.cell(row=i, column=6, value=s.tier_label)
        ws.cell(row=i, column=7, value=s.tier_rate_pct).number_format = _PCT_FMT
        ws.cell(row=i, column=8, value=s.commission_amount).number_format = _MONEY_FMT

    sa_count = len(summaries)
    if summaries:
        last = sa_count + 2
        ws.cell(row=last, column=1, value="SA TOTAL").font = Font(bold=True)
        ws.cell(row=last, column=2, value=sum(s.order_count for s in summaries)).font = Font(bold=True)
        for col, attr in [(3, "total_gross_sales"), (4, "total_net_sales"), (8, "commission_amount")]:
            cell = ws.cell(row=last, column=col, value=sum(getattr(s, attr) for s in summaries))
            cell.number_format = _MONEY_FMT
            cell.font = Font(bold=True)

    if house:
        # Visual gap, then a separate "House sales" row in italic.
        row = sa_count + 4
        ws.cell(row=row, column=1, value="House sales (COMPANY SALES — no commission)").font = Font(
            italic=True, bold=True
        )
        row += 1
        ws.cell(row=row, column=1, value="COMPANY SALES").font = Font(italic=True)
        ws.cell(row=row, column=2, value=house.order_count).font = Font(italic=True)
        c3 = ws.cell(row=row, column=3, value=house.total_gross_sales)
        c3.number_format = _MONEY_FMT
        c3.font = Font(italic=True)
        c4 = ws.cell(row=row, column=4, value=house.total_net_sales)
        c4.number_format = _MONEY_FMT
        c4.font = Font(italic=True)
        avg = round(house.total_gross_sales / house.order_count, 2) if house.order_count else 0.0
        c5 = ws.cell(row=row, column=5, value=avg)
        c5.number_format = _MONEY_FMT
        c5.font = Font(italic=True)
        ws.cell(row=row, column=6, value="House account — no commission").font = Font(italic=True)
        c8 = ws.cell(row=row, column=8, value=0.0)
        c8.number_format = _MONEY_FMT
        c8.font = Font(italic=True)

    _autosize(ws)


def _payment_str(order: OrderResult | None) -> str:
    """Compact payment summary, e.g. 'Mastercard Credit *2956, Cash'."""
    if order is None:
        return ""
    parts = []
    for p in order.parsed.payments:
        s = p.method.value.replace("_", " ").title()
        if p.last4:
            s += f" *{p.last4}"
        parts.append(s)
    return ", ".join(parts)


def _sa_order_data(c, order, tiers_cfg, tier_rate_pct) -> dict:
    """One order line for the SA sheet, with the real per-order commission.
    A partially-clearance order arrives as two contributions (normal +
    clearance); charges come from this portion's gross − net."""
    charges_share = round(c.gross_share - c.net_share, 2)
    is_clr = getattr(c, "is_clearance", False)
    flat_rule = tiers_cfg.flat_rule_for(order.channel) if order is not None else None
    if is_clr:
        commission = round(tiers_cfg.clearance_flat_amount * c.share_pct, 2)
    elif flat_rule is not None:
        commission = round(flat_rule.amount_per_order * c.share_pct, 2)
    else:
        commission = round(c.net_share * tier_rate_pct / 100.0, 2)
    # Effective charge rate on this SA's gross share (blends multiple payment
    # methods on split-payment orders). Charges in Excel = Gross × this rate.
    charge_rate = round(charges_share / c.gross_share, 6) if c.gross_share else 0.0
    return {
        "order": c.order_number,
        "date": c.order_date.strftime("%Y-%m-%d"),
        "channel": order.channel if order else "",
        "payment": _payment_str(order),
        "gross": c.gross_share,
        "discount": getattr(c, "discount_share", 0.0),
        "charge_rate": charge_rate,
        "charges": charges_share,
        "net": c.net_share,
        "share": c.share_pct,
        "commission": commission,
        "is_clr": is_clr,
    }


def _sa_section(ws: Worksheet, row: int, text: str, fill: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_SA_NC)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 20


def _sa_theader(ws: Worksheet, row: int) -> None:
    for col, h in enumerate(_SA_HEADERS, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.border = _BORDER
        c.alignment = Alignment(
            horizontal="right" if col in _SA_RIGHT_COLS else "left",
            vertical="center",
        )
    ws.row_dimensions[row].height = 18


def _sa_order_row(
    ws: Worksheet, row: int, d: dict, rate: float, flat: float,
    zebra_fill: str | None, *, blank_commission: bool = False,
) -> None:
    """One order line. Charge %, Charges, Net and Commission are live formulas:
    Charges = Gross × Charge%;  Net = Gross − Charges;
    Commission = Net × tier-rate (flat for clearance, blank for reference)."""
    label = d["order"] + ("  ·clearance" if d["is_clr"] else "")
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=d["date"])
    ws.cell(row=row, column=3, value=d["channel"])
    ws.cell(row=row, column=4, value=d["payment"])
    ws.cell(row=row, column=5, value=d["gross"])                  # E: value
    ws.cell(row=row, column=6, value=d["discount"])              # F: discount (value)
    ws.cell(row=row, column=7, value=d["charge_rate"])            # G: charge % (value)
    ws.cell(row=row, column=8, value=f"=E{row}*G{row}")           # H: charges = gross×rate
    ws.cell(row=row, column=9, value=f"=E{row}-H{row}")           # I: net = gross−charges
    ws.cell(row=row, column=10, value=d["share"])                 # J: share % (value)
    if blank_commission:
        ws.cell(row=row, column=11, value=None)                   # reference row
    elif d["is_clr"]:
        # K: flat RM per ORDER, split by this SA's share (J). A 30% share on a
        # clearance order earns 30% × RM10 = RM3 — matching the engine and the
        # on-screen breakdown, not the full flat.
        ws.cell(row=row, column=11, value=f"=J{row}*{flat}")
    else:
        # K: net × the LIVE tier rate cell, so deleting/adding rows re-picks
        # the bracket and the whole sheet still ties.
        ws.cell(row=row, column=11, value=f"=I{row}*{_TIER_RATE_REF}/100")
    for col in range(1, _SA_NC + 1):
        c = ws.cell(row=row, column=col)
        c.border = _BORDER
        if zebra_fill:
            c.fill = PatternFill("solid", fgColor=zebra_fill)
        if col in _SA_NUM_COLS:
            c.number_format = _MONEY_FMT
        elif col == 7:
            c.number_format = "0.00%"
        elif col == 10:
            c.number_format = "0%"
        if col in _SA_RIGHT_COLS:
            c.alignment = Alignment(horizontal="right")


def _sa_subtotal(
    ws: Worksheet, row: int, label: str, start: int, end: int,
    *, with_commission: bool = True,
) -> None:
    """Subtotal row with =SUM() formulas (Charges, Net, Commission)."""
    for col in range(1, _SA_NC + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=_SUBTOTAL_FILL)
        cell.border = _BORDER
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True, color=_INK)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    lc.alignment = Alignment(horizontal="right", indent=1)

    def _put(col: int, formula) -> None:
        cell = ws.cell(row=row, column=col, value=formula)
        cell.number_format = _MONEY_FMT
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

    _put(8, f"=SUM(H{start}:H{end})" if end >= start else 0)   # Charges
    _put(9, f"=SUM(I{start}:I{end})" if end >= start else 0)   # Net
    if with_commission:
        _put(11, f"=SUM(K{start}:K{end})" if end >= start else 0)  # Commission


def _sa_refund_table(ws: Worksheet, row: int, refunds: list) -> tuple[int, str]:
    """Refunds table: refunded orders for this SA with a manual Clawback column
    that deducts from the total. Returns (next_row, clawback_subtotal_cell)."""
    _sa_section(
        ws, row,
        "REFUNDS — COMMISSION CLAWBACK  ·  enter the commission to deduct "
        "(as a negative) if it was already paid",
        "B42318",
    )
    row += 1
    hdrs = {1: "Order #", 2: "Order date", 3: "Settled", 4: "Payment",
            5: "Refunded amount", 10: "Share %", 11: "Clawback (enter −)"}
    for col in range(1, _SA_NC + 1):
        c = ws.cell(row=row, column=col, value=hdrs.get(col))
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.border = _BORDER
        c.alignment = Alignment(
            horizontal="right" if col in (5, 10, 11) else "left", vertical="center"
        )
    ws.row_dimensions[row].height = 18
    row += 1
    start = row
    for i, (o, share) in enumerate(refunds):
        settled = (o.settlement_date or o.order_date).strftime("%Y-%m")
        vals = {
            1: o.order_number,
            2: o.order_date.strftime("%Y-%m-%d"),
            3: settled,
            4: _payment_str(o),
            5: round(o.gross_total * share, 2),
            10: share,
            11: None,  # manual clawback entry
        }
        fill = "FBECEA" if i % 2 else None
        for col in range(1, _SA_NC + 1):
            c = ws.cell(row=row, column=col, value=vals.get(col))
            c.border = _BORDER
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if col == 5 or col == 11:
                c.number_format = _MONEY_FMT
            elif col == 10:
                c.number_format = "0%"
            if col in (5, 10, 11):
                c.alignment = Alignment(horizontal="right")
        row += 1
    for col in range(1, _SA_NC + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=_SUBTOTAL_FILL)
        cell.border = _BORDER
    lc = ws.cell(row=row, column=1, value="TOTAL CLAWBACK (deducted from commission)")
    lc.font = Font(bold=True, color="B42318")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    lc.alignment = Alignment(horizontal="right", indent=1)
    cc = ws.cell(row=row, column=11, value=f"=SUM(K{start}:K{row - 1})" if row - 1 >= start else 0)
    cc.number_format = _MONEY_FMT
    cc.font = Font(bold=True, color="B42318")
    cc.alignment = Alignment(horizontal="right")
    return row + 2, f"K{row}"


def _sa_order_table(
    ws: Worksheet, row: int, section: str, section_fill: str, zebra: str,
    data: list[dict], rate: float, flat: float, label: str,
) -> tuple[int, int]:
    """Render a section header + table of orders + subtotal. Returns
    (next_row, subtotal_row)."""
    _sa_section(ws, row, section, section_fill)
    row += 1
    _sa_theader(ws, row)
    row += 1
    start = row
    for i, d in enumerate(data):
        _sa_order_row(ws, row, d, rate, flat, zebra if i % 2 else None)
        row += 1
    _sa_subtotal(ws, row, label, start, row - 1)
    return row + 2, row


def _sa_manual_prev_table(
    ws: Worksheet, row: int, n_rows: int = 6,
) -> tuple[int, str, str]:
    """Blank, manually-filled table for previous-month sales the system did not
    auto-capture, so they are never missed. The user types Order #, Date,
    Channel, Payment, Gross, Charge % and Share %; Charges, Net and Commission
    calculate live (using the same tier rate) and the subtotal rolls into the
    totals. Returns (next_row, commission_subtotal_cell, net_subtotal_cell)."""
    _sa_section(
        ws, row,
        "PREVIOUS-MONTH SALES MISSED  ·  enter manually — Charges, Net & "
        "Commission fill in automatically",
        _MUTED,
    )
    row += 1
    _sa_theader(ws, row)
    row += 1
    start = row
    for i in range(n_rows):
        zebra = _ZEBRA if i % 2 else None
        for col in range(1, _SA_NC + 1):
            c = ws.cell(row=row, column=col)
            c.border = _BORDER
            if zebra:
                c.fill = PatternFill("solid", fgColor=zebra)
            if col in _SA_RIGHT_COLS:
                c.alignment = Alignment(horizontal="right")
        # Live formulas — blank until the row's Gross is filled in.
        ws.cell(row=row, column=8, value=f'=IF(E{row}="","",E{row}*G{row})')
        ws.cell(row=row, column=9, value=f'=IF(E{row}="","",E{row}-H{row})')
        ws.cell(
            row=row, column=11,
            value=f'=IF(E{row}="","",I{row}*{_TIER_RATE_REF}/100)',
        )
        for col in (5, 6, 8, 9, 11):
            ws.cell(row=row, column=col).number_format = _MONEY_FMT
        ws.cell(row=row, column=7).number_format = "0.00%"
        ws.cell(row=row, column=10).number_format = "0%"
        row += 1
    _sa_subtotal(ws, row, "PREVIOUS-MONTH MISSED TOTAL", start, row - 1)
    return row + 2, f"K{row}", f"I{row}"


def _sa_bonus_table(ws: Worksheet, row: int, sa: SACommission, achieved_formula: str) -> tuple[int, str]:
    """Overachievement-bonus summary with live formulas. Returns
    (next_row, cash_incentive_cell)."""
    _sa_section(
        ws, row,
        f"OVERACHIEVEMENT BONUS  ·  {sa.bonus_season} season"
        f"  ·  RM500 per full RM50,000 over target",
        "6D28D9",
    )
    row += 1
    tgt_row = row
    specs = [
        ("Monthly sales target", sa.bonus_target, _MONEY_FMT),
        ("Sales achieved (excl. clearance)", achieved_formula, _MONEY_FMT),
        ("Extra above target", f"=MAX(0,K{tgt_row + 1}-K{tgt_row})", _MONEY_FMT),
        ("Tiers achieved (× RM50,000)", f"=INT(K{tgt_row + 2}/50000)", "0"),
        ("Cash incentive (× RM500)", f"=K{tgt_row + 3}*500", _MONEY_FMT),
    ]
    for i, (label, val, fmt) in enumerate(specs):
        is_cash = i == len(specs) - 1
        for col in range(1, _SA_NC + 1):
            c = ws.cell(row=row, column=col)
            c.border = _BORDER
            if is_cash:
                c.fill = PatternFill("solid", fgColor="EDE9FE")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        lc = ws.cell(row=row, column=1, value=label)
        lc.alignment = Alignment(horizontal="right", indent=1)
        lc.font = Font(bold=is_cash, color=_INK)
        vc = ws.cell(row=row, column=11, value=val)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="right")
        vc.font = Font(bold=is_cash, color="6D28D9" if is_cash else _INK)
        row += 1
    return row + 1, f"K{tgt_row + 4}"


def _sa_signoff(ws: Worksheet, row: int) -> int:
    """Payout sign-off block: a verification statement plus Prepared / Verified /
    Approved signature lines with name and date. Signing it marks the commission
    on this sheet as checked and cleared for payout. Returns the next row."""
    _sa_section(ws, row, "PAYOUT SIGN-OFF", _INK)
    row += 1
    stmt = ws.cell(
        row=row, column=1,
        value="I confirm the TOTAL COMMISSION above has been reviewed, verified "
              "correct, and is approved for payout.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_SA_NC)
    stmt.font = Font(italic=True, color=_INK)
    stmt.alignment = Alignment(horizontal="left", indent=1, vertical="center")
    ws.row_dimensions[row].height = 18
    row += 2  # a spacer row above the signature lines

    line_border = Border(bottom=Side(style="thin", color=_INK))
    blocks = [(1, 3, "Prepared by"), (5, 7, "Verified by"), (9, 11, "Approved by")]
    sign_row, role_row, name_row, date_row = row, row + 1, row + 2, row + 3
    ws.row_dimensions[sign_row].height = 34  # room for a physical signature
    for c1, c2, label in blocks:
        for c in range(c1, c2 + 1):
            ws.cell(row=sign_row, column=c).border = line_border
        ws.merge_cells(start_row=sign_row, start_column=c1, end_row=sign_row, end_column=c2)
        rl = ws.cell(row=role_row, column=c1, value=label)
        rl.font = Font(bold=True, color=_INK)
        ws.cell(row=name_row, column=c1, value="Name:").font = Font(size=9, color=_MUTED)
        ws.cell(row=date_row, column=c1, value="Date:").font = Font(size=9, color=_MUTED)
    return date_row + 1


def _build_sa_sheet(
    ws: Worksheet,
    sa: SACommission,
    orders_by_number: dict[str, OrderResult],
    tiers_cfg: TiersConfig,
    *,
    payout_month: str | None = None,
    payout_label: str | None = None,
    refunds: list | None = None,
) -> None:
    ws.title = f"SA - {sa.sa_name}"[:31]  # Excel sheet name limit
    ws.sheet_view.showGridLines = False
    rate = sa.tier_rate_pct
    flat = tiers_cfg.clearance_flat_amount
    plabel = payout_label or "current month"

    # ---- Title block -------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_SA_NC)
    title = ws.cell(row=1, column=1, value=sa.sa_name)
    title.font = Font(bold=True, size=18, color=_INK)
    title.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_SA_NC)
    sub = ws.cell(
        row=2, column=1,
        value=f"{plabel}     •     Figures recalculate live — add or delete order "
              f"rows and the tier rate (top-right, cell M3) and all totals re-tie.",
    )
    sub.font = Font(size=10, italic=True, color=_MUTED)
    sub.alignment = Alignment(vertical="center", indent=1)

    # ---- Classify orders ---------------------------------------------------
    rows = [
        _sa_order_data(c, orders_by_number.get(c.order_number), tiers_cfg, rate)
        for c in sorted(sa.contributions, key=lambda c: c.order_date)
    ]
    clearance = [d for d in rows if d["is_clr"]]
    non_clr = [d for d in rows if not d["is_clr"]]
    # Split paid orders by when they were ORDERED vs the payout month.
    if payout_month:
        prev = [d for d in non_clr if d["date"][:7] < payout_month]
        current = [d for d in non_clr if d["date"][:7] >= payout_month]
    else:
        prev, current = [], non_clr

    row = 4
    subtotal_cells: list[str] = []  # commission subtotal cells to sum

    # ---- Table 1: previous-month sales paid in current month ---------------
    if prev:
        row, sub_r = _sa_order_table(
            ws, row,
            f"PREVIOUS-MONTH SALES PAID IN {plabel.upper()}",
            _MUTED, _ZEBRA, prev, rate, flat, "PREVIOUS-MONTH TOTAL",
        )
        subtotal_cells.append(f"K{sub_r}")   # commission subtotal
        t1_net_cell = f"I{sub_r}"            # net subtotal
    else:
        t1_net_cell = None

    # ---- Table 1b: manual catch-up for missed previous-month sales ---------
    row, missed_comm_cell, missed_net_cell = _sa_manual_prev_table(ws, row)
    subtotal_cells.append(missed_comm_cell)

    # ---- Table 2: current-month sales completed ----------------------------
    row, sub_r = _sa_order_table(
        ws, row, f"{plabel.upper()} SALES COMPLETED", _NAVY, _ZEBRA,
        current, rate, flat, f"{plabel.upper()} SALES TOTAL",
    )
    subtotal_cells.append(f"K{sub_r}")
    t2_net_cell = f"I{sub_r}"

    # ---- Table 3: clearance sales ------------------------------------------
    if clearance:
        row, sub_r = _sa_order_table(
            ws, row,
            f"CLEARANCE SALES  ·  flat RM{flat:,.0f} per order"
            f"  ·  excluded from sales total & tier",
            _GOLD, _CLR_ZEBRA, clearance, rate, flat, "CLEARANCE TOTAL",
        )
        subtotal_cells.append(f"K{sub_r}")

    # ---- Table 4: overachievement bonus (SAs with a scheme) ----------------
    bonus_cell = None
    if sa.bonus_season:
        row, bonus_cell = _sa_bonus_table(ws, row, sa, f"={_TIER_NET_CELL}")

    # ---- Table 5: refunds / commission clawback ----------------------------
    clawback_cell = None
    if refunds:
        row, clawback_cell = _sa_refund_table(ws, row, refunds)

    # Commission before bonus = tier + clearance commissions (+ any clawback).
    before_formula = "=" + "+".join(subtotal_cells) + (f"+{clawback_cell}" if clawback_cell else "")

    # ---- "Before bonus" line (only when there's a bonus to add on top) ------
    if bonus_cell:
        for col in range(1, _SA_NC + 1):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor="D6E4E1")
            c.border = _BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        lbl = ws.cell(row=row, column=1, value="Commission before overachievement bonus")
        lbl.font = Font(bold=True, color=_INK)
        lbl.alignment = Alignment(horizontal="right", indent=1)
        bcell = ws.cell(row=row, column=11, value=before_formula)
        bcell.number_format = _MONEY_FMT
        bcell.font = Font(bold=True)
        bcell.alignment = Alignment(horizontal="right")
        grand = f"=K{row}+{bonus_cell}"
        row += 1
    else:
        grand = before_formula

    # ---- Grand total commission -------------------------------------------
    for col in range(1, _SA_NC + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=_TEAL)
        ws.cell(row=row, column=col).border = _BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    g = ws.cell(row=row, column=1, value="TOTAL COMMISSION")
    g.font = Font(bold=True, size=12, color="FFFFFF")
    g.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    gc = ws.cell(row=row, column=11, value=grand)
    gc.number_format = _MONEY_FMT
    gc.font = Font(bold=True, size=12, color="FFFFFF")
    gc.alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 22

    # ---- Monthly Google reviews (manual entry) -----------------------------
    gr_row = row + 2
    lbl = ws.cell(
        row=gr_row, column=1,
        value="Total Google reviews collected this month:",
    )
    ws.merge_cells(start_row=gr_row, start_column=1, end_row=gr_row, end_column=7)
    lbl.font = Font(bold=True, color=_INK)
    lbl.alignment = Alignment(horizontal="right", indent=1, vertical="center")
    ws.merge_cells(start_row=gr_row, start_column=8, end_row=gr_row, end_column=11)
    for c in range(8, _SA_NC + 1):
        cell = ws.cell(row=gr_row, column=c)
        cell.border = _BORDER
        cell.fill = PatternFill("solid", fgColor="FFF7E6")  # fill-in box
    ws.cell(row=gr_row, column=8).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.row_dimensions[gr_row].height = 20

    # ---- Payout sign-off ---------------------------------------------------
    _sa_signoff(ws, gr_row + 2)

    # ---- Live tier helper (top-right: M2 = sales net, M3 = tier rate) -------
    kh = ws.cell(row=1, column=12, value="LIVE TIER")
    kh.font = Font(bold=True, size=10, color="1F4E79")
    ws.cell(row=2, column=12, value="Net (excl. clearance)").font = Font(size=9, color=_MUTED)
    ws.cell(row=3, column=12, value="Tier rate applied").font = Font(size=9, color=_MUTED)
    net_refs = [c for c in [t1_net_cell, t2_net_cell, missed_net_cell] if c]
    nc = ws.cell(row=2, column=13, value=("=" + "+".join(net_refs)) if net_refs else 0)
    nc.number_format = _MONEY_FMT
    nc.font = Font(bold=True)
    rc = ws.cell(row=3, column=13, value=_tier_rate_formula(_TIER_NET_CELL, tiers_cfg.tiers))
    rc.number_format = '0.0"%"'
    rc.font = Font(bold=True, color="1F4E79")

    # ---- Column widths & freeze -------------------------------------------
    for col, w in enumerate([16, 11, 12, 30, 13, 11, 9, 12, 13, 8, 13], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 12
    ws.freeze_panes = "A4"


def _build_review_sheet(ws: Worksheet, orders: list[OrderResult]) -> None:
    ws.title = "Review log"
    review = [o for o in orders if o.needs_review]
    headers = ["Order #", "Date", "Gross", "Channel", "Note", "Flags"]
    _write_header(ws, headers)
    for i, o in enumerate(review, start=2):
        ws.cell(row=i, column=1, value=o.order_number)
        ws.cell(row=i, column=2, value=o.order_date.strftime("%Y-%m-%d"))
        ws.cell(row=i, column=3, value=o.gross_total).number_format = _MONEY_FMT
        ws.cell(row=i, column=4, value=o.channel)
        ws.cell(row=i, column=5, value=o.parsed.raw_note)
        ws.cell(row=i, column=6, value=" | ".join(o.parsed.review_flags))
    _autosize(ws, max_w=80)


def _build_excluded_sheet(ws: Worksheet, orders: list[OrderResult]) -> None:
    ws.title = "Excluded"
    excluded = [o for o in orders if o.excluded]
    headers = ["Order #", "Date", "Gross", "Channel", "Order Status", "Financial Status", "Reason"]
    _write_header(ws, headers)
    for i, o in enumerate(excluded, start=2):
        ws.cell(row=i, column=1, value=o.order_number)
        ws.cell(row=i, column=2, value=o.order_date.strftime("%Y-%m-%d"))
        ws.cell(row=i, column=3, value=o.gross_total).number_format = _MONEY_FMT
        ws.cell(row=i, column=4, value=o.channel)
        ws.cell(row=i, column=5, value=o.order_status)
        ws.cell(row=i, column=6, value=o.financial_status)
        ws.cell(row=i, column=7, value=o.excluded_reason or "")
    _autosize(ws)


def _build_settings_sheet(ws: Worksheet, settings: AppSettings) -> None:
    ws.title = "Settings snapshot"
    row = 1
    ws.cell(row=row, column=1, value="Sales Advisors").font = Font(bold=True, size=12)
    row += 1
    for sa in settings.sa_list.sas:
        ws.cell(row=row, column=1, value=sa.name)
        ws.cell(row=row, column=2, value="active" if sa.active else "inactive")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Tiers").font = Font(bold=True, size=12)
    row += 1
    for t in settings.tiers.tiers:
        ws.cell(row=row, column=1, value=f"RM {t.min_net:,.2f}")
        ws.cell(row=row, column=2, value=f"RM {t.max_net:,.2f}" if t.max_net is not None else "and above")
        ws.cell(row=row, column=3, value=f"{t.rate_pct}%")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Channel flat commissions").font = Font(bold=True, size=12)
    row += 1
    for r in settings.tiers.channel_flat_commissions:
        ws.cell(row=row, column=1, value=r.channel)
        ws.cell(row=row, column=2, value=r.amount_per_order).number_format = _MONEY_FMT
        ws.cell(row=row, column=3, value=r.label)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Active rate version").font = Font(bold=True, size=12)
    row += 1
    if settings.rates.versions:
        v = max(settings.rates.versions, key=lambda v: v.effective_from)
        ws.cell(row=row, column=1, value=f"Effective from {v.effective_from.isoformat()}")
        row += 1
        ws.cell(row=row, column=1, value="SenangPay card").font = Font(italic=True)
        ws.cell(row=row, column=2, value=v.senangpay_card_pct).number_format = _PCT_FMT
        row += 1
        ws.cell(row=row, column=1, value="SenangPay FPX").font = Font(italic=True)
        ws.cell(row=row, column=2, value=v.senangpay_fpx_pct).number_format = _PCT_FMT
        row += 1
        for rate in v.rates:
            ws.cell(row=row, column=1, value=rate.label)
            if rate.rate_pct is not None:
                ws.cell(row=row, column=2, value=rate.rate_pct).number_format = _PCT_FMT
            else:
                ws.cell(row=row, column=2, value="(not configured)")
            row += 1
    _autosize(ws)


def _build_house_sheet(
    ws: Worksheet,
    house: HouseSalesSummary,
    orders_by_number: dict[str, OrderResult],
) -> None:
    """Audit-trail sheet for COMPANY SALES (no commission column)."""
    ws.title = "House - COMPANY SALES"[:31]
    headers = [
        "Order #",
        "Date",
        "Channel",
        "Order Gross",
        "Payment Method",
        "Last 4",
        "Portion Gross",
        "Rate Applied",
        "Charge Amount",
        "Portion Net",
        "Share %",
        "Net to House",
    ]
    _write_header(ws, headers)
    row = 2
    for c in house.contributions:
        order = orders_by_number.get(c.order_number)
        if order is None or order.excluded or not order.charges:
            ws.cell(row=row, column=1, value=c.order_number)
            ws.cell(row=row, column=2, value=c.order_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=order.channel if order else "")
            ws.cell(row=row, column=4, value=order.gross_total if order else 0).number_format = _MONEY_FMT
            ws.cell(row=row, column=11, value=c.share_pct).number_format = "0.0%"
            ws.cell(row=row, column=12, value=c.net_share).number_format = _MONEY_FMT
            row += 1
            continue
        for ch in order.charges:
            ws.cell(row=row, column=1, value=order.order_number)
            ws.cell(row=row, column=2, value=order.order_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=order.channel)
            ws.cell(row=row, column=4, value=order.gross_total).number_format = _MONEY_FMT
            ws.cell(row=row, column=5, value=ch.method.value)
            ws.cell(row=row, column=6, value=ch.last4 or "")
            ws.cell(row=row, column=7, value=ch.gross).number_format = _MONEY_FMT
            ws.cell(row=row, column=8, value=ch.rate_label)
            ws.cell(row=row, column=9, value=ch.charge).number_format = _MONEY_FMT
            ws.cell(row=row, column=10, value=ch.net).number_format = _MONEY_FMT
            ws.cell(row=row, column=11, value=c.share_pct).number_format = "0.0%"
            if ch is order.charges[0]:
                ws.cell(row=row, column=12, value=c.net_share).number_format = _MONEY_FMT
            row += 1
    ws.cell(row=row, column=1, value="TOTAL HOUSE NET").font = Font(bold=True)
    ws.cell(row=row, column=12, value=house.total_net_sales).number_format = _MONEY_FMT
    ws.cell(row=row, column=12).font = Font(bold=True)
    _autosize(ws)


def build_workbook(
    orders: list[OrderResult],
    report: CommissionReport,
    settings: AppSettings,
    *,
    payout_month: str | None = None,
    payout_label: str | None = None,
    refunded_orders: list[OrderResult] | None = None,
    all_orders: list[OrderResult] | None = None,
) -> bytes:
    """Render the full report and return raw .xlsx bytes.

    `orders` are the payout-month's kept orders (used for the SA sheets).
    `all_orders` is the full uploaded set — used for the Review and Excluded
    sheets so held-back orders (COD/unpaid/refunded) are visible and never
    silently disappear. Defaults to `orders` if not given.
    """
    audit = all_orders if all_orders is not None else orders
    wb = Workbook()
    summary_ws = wb.active
    _build_summary_sheet(summary_ws, report.sa_summaries, report.house)

    # Group refunded orders by the SA(s) named on them.
    refunds_by_sa: dict[str, list] = {}
    for o in (refunded_orders or []):
        for sh in o.parsed.sa_shares:
            if sh.name == HOUSE_ACCOUNT:
                continue
            refunds_by_sa.setdefault(sh.name, []).append((o, sh.share))

    by_number = {o.order_number: o for o in orders}
    for s in report.sa_summaries:
        ws = wb.create_sheet()
        _build_sa_sheet(
            ws, s, by_number, settings.tiers,
            payout_month=payout_month, payout_label=payout_label,
            refunds=refunds_by_sa.get(s.sa_name),
        )

    if report.house:
        _build_house_sheet(wb.create_sheet(), report.house, by_number)

    _build_review_sheet(wb.create_sheet(), audit)
    _build_excluded_sheet(wb.create_sheet(), audit)
    _build_settings_sheet(wb.create_sheet(), settings)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
